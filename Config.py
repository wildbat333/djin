import pandas as pd
from openpyxl import Workbook

file_name_input = "smart_ivanovo.xlsx"
file_name_input_2 = "sv_k.xlsx"
file_name_input_3 = "sv_e.xlsx"
file_name_result = "smart_ivanovo_result.xlsx"

d = ['0D1-035', 'Семечки "Джинн"35г.(100)', 100,
     '0D1-070', 'Семечки "Джинн"70г.(50)', 50,
     '0D1-140', 'Семечки "Джинн"(10%)140г.(25)', 25,
     '0D1-250', 'Семечки "Джинн"250г.(14)', 14,
     '0D1-350', 'Семечки "Джинн"350г.(10)', 10]
d_s = ['0D2-035', 'Семечки "Джинн"(соленые)35г.(100)', 100,
       '0D2-070', 'Семечки "Джинн"(соленые)70г.(50)', 50,
       '0D2-140', 'Семечки "Джинн"(соленые)(10%)140г.(25)', 25,
       '0D2-250', 'Семечки "Джинн"(соленые)250г.(14)', 14,
       '0D2-350', 'Семечки "Джинн"(соленые)350г.(10)', 10]
v = ['0V1-050', 'Семечки "Солнечный Великан"Джинн"50г.(60)', 60,
     '0V1-100', 'Семечки "Солнечный Великан"Джинн"100г.(30)', 30,
     '0V1-120', 'Семечки "Солнечный Великан"Джинн"120г.(25)', 25,
     '0V1-200', 'Семечки "Солнечный Великан"Джинн"200г.(15)', 15]
v_s = ['0V2-050', 'Семечки "Солнечный Великан"Джинн"(соленые)50г.(60)', 60,
       '0V2-100', 'Семечки "Солнечный Великан"Джинн"(соленые)100г.(30', 30,
       '0V2-120', 'Семечки "Солнечный Великан"Джинн"(соленые)120г.(25', 25,
       '0V2-200', 'Семечки "Солнечный Великан"Джинн"(соленые)200г.(15', 15,
       '0V2-300', 'Семечки "Солнечный Великан"Джинн"(соленые)300г.(10', 10]
v_o_s = ['0V3-100', 'Семечки "Солнечный Великан"Джинн"(особо сол100г(30', 30]
m = ['0M1-070', 'Семечки "Мастер Жарки"70г.(50)', 50,
     '0M1-140', 'Семечки "Мастер Жарки"140г.(25)', 25,
     '0M1-250', 'Семечки "Мастер Жарки"250г.(14)', 14,
     '0M1-350', 'Семечки "Мастер Жарки"350г.(10)', 10]
k = ['0K2-050', 'Арахис "Караван орехов"50г.(40)', 40,
     '0K2-090', 'Арахис "Караван орехов"90г.(22)', 22,
     '0K2-150', 'Арахис "Караван орехов"150г.(20)', 13.3,
     '0K2-500', 'Арахис "Караван орехов"500г.(50)', 4]
k_s = ['0S2-040', 'Арахис "Караван орехов"Стандарт"40г.(50)', 50,
       '0S2-090', 'Арахис "Караван орехов"Стандарт"90г.(22)', 22]
t = ['0T2-050', 'Семечки тыквы "Джинн"(соленые)50г.(70)', 70,
     '0T2-100', 'Семечки тыквы "Джинн"(соленые)100г.(35)', 35]

#Dolgacheva = ['ПВЗ Шуя 1', 'ПВЗ Шуя 2', 'ПВЗ Родники 1']  ##Долгачева Ирина (ЭК СМАРТ
Drozdova = ['ПВЗ Шуя 1', 'ПВЗ Шуя 2', 'ПВЗ Родники 1', 'ПВЗ Вичуга 1']  ##Дроздова Марина (ЭК СМАРТ
# добавлено ['ПВЗ Шуя 1', 'ПВЗ Шуя 2', 'ПВЗ Родники 1']
Korsakova = ['ПВЗ Воскресенское 1', 'ПВЗ Новые горки 1', 'ПВЗ Лежнево 1', 'ПВЗ Лежнево 2', 'ПВЗ Шилыково 1']  ##Корсакова Елена (ЭК СМАРТ
Krasnova = ['ПВЗ Жуково 1', 'ПВЗ Иваново Благова', 'ПВЗ Иваново Дружба']  ##Краснова Наталья (ЭК СМАР
Shutova = ['ПВЗ Васильевское 1', 'ПВЗ Воскресенское 2', 'ПВЗ Колобово 1', 'ПВЗ Михалково 1', 'ПВЗ Палех 1',
           'ПВЗ Савино 1', 'ПВЗ Центральный 1']  ##Шутова Ольга (ЭК СМАРТ)
Migushova = ['ПВЗ Иваново Академическая', 'ПВЗ Иваново Демидова', 'ПВЗ Иваново Бубнова', 'ПВЗ Иваново Сакко',
             'ПВЗ Иваново Самойлова', 'ПВЗ Суворова', 'ПВЗ Иваново 9 Января', 'ПВЗ Иваново Победы',
             'ПВЗ Иваново Слесарный', 'ПВЗ Иваново Афанасьева']  ##Мигушова Надежда (ЭК СМАР
Fedotova = ['ПВЗ Михалево 1', 'ПВЗ НовоТалицы 2', 'ПВЗ НовоТалицы 3', 'ПВЗ Тейково 2', 'ПВЗ Чернореченский 1',
            'ПВЗ Чернореченский 2']  ##Федотова Анна (ЭК СМАРТ)
Chepa = ['ПВЗ Авдотьино 1', 'ПВЗ Авдотьино 2', 'ПВЗ Иваново Гарнизон', 'ПВЗ Иваново Дюковская', 'ПВЗ Иваново Н.Неман',
         'ПВЗ Приволжск 1', 'ПВЗ Приволжск 2', 'ПВЗ Приволжск 3']  ##Чепа Елена (ЭК СМАРТ)
Ogurova = ['ПВЗ Заволжск 1', 'ПВЗ Наволоки 1', 'ПВЗ Кинешма 1', 'ПВЗ Кинешма 2', 'ПВЗ Кинешма 3']##Огурова Ольга (ЭК СМАРТ)
# добавлено 'ПВЗ Наволоки 1', 'ПВЗ Кинешма 1', 'ПВЗ Кинешма 2'

tp_1c = ['Дроздова Марина (ЭК СМАРТ', 'Корсакова Елена (ЭК СМАРТ',
         'Огурова Ольга (ЭК СМАРТ)', 'Федотова Анна (ЭК СМАРТ)', 'Краснова Наталья (ЭК СМАР',
         'Мигушова Надежда (ЭК СМАР', 'Чепа Елена (ЭК СМАРТ)', 'Шутова Ольга (ЭК СМАРТ)'] #убрана Долгачева
tp = ['Дроздова Марина', 'Корсакова Елена', 'Огурова Ольга', 'Федотова Анна',
      'Краснова Наталья', 'Мигушова Надежда', 'Чепа Елена', 'Шутова Ольга'] #убрана Долгачева

#Количества товаров в группах отчета СВ и общая их сумма
number_of_products_in_group_1 = 5
number_of_products_in_group_2 = 2
number_of_products_in_group_3 = 3
number_of_products = number_of_products_in_group_1 + number_of_products_in_group_2 + number_of_products_in_group_3

row_with_itogo = 12

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet('Сводка 2.0')
ws2 = wb.create_sheet('Корректировка')
ws3 = wb.create_sheet('Сводка')
ws4 = wb.create_sheet('Пачки')
ws5 = wb.create_sheet('Полная')
ws6 = wb.create_sheet('СВ')
wb.remove(ws)
wb.save(file_name_result)


df = pd.read_excel(file_name_input, header=None)
df_2 = pd.read_excel(file_name_input_2, header=None)

cel_k = []
cel_e = []

itog_k = []
itog_e = []


clmns = list(df_2)
#print(clmns)
for c in range(len(clmns)):     #Поиск заголовка Итого в 12 строке - шапка
    if df_2.loc[row_with_itogo, c] == 'Итого': cL_itogo = c
for g, row in df_2.iterrows():
   # if df_2.loc[g, 0] == 'Долгачева Ирина': cel_k.append(g)
    if df_2.loc[g, 0] == tp[0]: cel_k.append(g)
    if df_2.loc[g, 0] == tp[1]: cel_k.append(g)
    if df_2.loc[g, 0] == tp[2]: cel_k.append(g)
    if df_2.loc[g, 0] == tp[3]: cel_k.append(g)

#Проверка не соскочил ли Итого
if df_2.loc[row_with_itogo, cL_itogo] == 'Итого':
    print('Количество строк не изменилось, Итого на месте!')
else:
    print('Количество строк изменилось! Сейчас в ячейке Итого: ', df_2.loc[row_with_itogo, cL_itogo],
          '\nИзмени номер строки!')

#Можно проверить какие ТП забираются из отчета СВ
# tp_from_sv_k = []
# for i in range(len(cel_k)):
#     tp_from_sv_k.append(df_2.loc[cel_k[i], 0])
#print(tp_from_sv_k)

# Отбрасываем строки Групп из таблицы по количествам товаров из констант сверху
for n in range(len(cel_k)):
    for i in range(number_of_products+2):
        if i == number_of_products_in_group_1:
            continue
        elif i == number_of_products_in_group_1 + number_of_products_in_group_2 + 1:
            continue
        elif i == number_of_products + 2:
            break
        else:
            itog_k.append(cel_k[n]+(i+1))


#Можно проверить какие позиции забираются из отчета СВ
products_groups_sv_k = []
for i in range(len(itog_k)):
    products_groups_sv_k.append(df_2.loc[itog_k[i], 0])
# print(products_groups_sv_k)

# Формируем список товаров для заголовков (полный)
columns = []
columns.extend(products_groups_sv_k[:number_of_products])

tp0 = []
tp0.append(tp[0])
tp1 = []
tp1.append(tp[1])
tp2 = []
tp2.append(tp[2])
tp3 = []
tp3.append(tp[3])
for i in range(len(itog_k)):
    if i < number_of_products:
        tp0.append(df_2.loc[itog_k[i], cL_itogo])
    elif i < number_of_products * 2:
        tp1.append(df_2.loc[itog_k[i], cL_itogo])
    elif i < number_of_products * 3:
        tp2.append(df_2.loc[itog_k[i], cL_itogo])
    elif i < number_of_products * 4:                #Это если 4 ТП иначе уменьшать или увеличивать
        tp3.append(df_2.loc[itog_k[i], cL_itogo])

df_3 = pd.read_excel(file_name_input_3, header=None)


clmns = list(df_3)
#print(clmns)

for c in range(len(clmns)):     #Поиск заголовка Итого в 12 строке - шапка
    if df_3.loc[row_with_itogo, c] == 'Итого': cL_itogo = c
for g, row in df_3.iterrows():
    if df_3.loc[g, 0] == tp[4]: cel_e.append(g)
    if df_3.loc[g, 0] == tp[5]: cel_e.append(g)
    if df_3.loc[g, 0] == tp[6]: cel_e.append(g)
    if df_3.loc[g, 0] == tp[7]: cel_e.append(g)

#Проверка не соскочил ли Итого
#print(df_3.loc[row_with_itogo, cL_itogo])  ##Итого

#Можно проверить какие ТП забираются из отчета СВ
# tp_from_sv_e = []
# for i in range(len(cel_e)):
#     tp_from_sv_e.append(df_2.loc[cel_e[i], 0])
#print(tp_from_sv_e)

# Отбрасываем строки Групп из таблицы по количествам товаров из констант сверху
for n in range(len(cel_e)):
    for i in range(number_of_products+2):
        if i == number_of_products_in_group_1:
            continue
        elif i == number_of_products_in_group_1 + number_of_products_in_group_2 + 1:
            continue
        elif i == number_of_products + 2:
            break
        else:
            itog_e.append(cel_e[n]+(i+1))

#Можно проверить какие позиции забираются из отчета СВ
# products_groups_sv_e = []
# for i in range(len(itog_e)):
#     products_groups_sv_e.append(df_3.loc[itog_e[i], 0])
# # print(products_groups_sv_e)

tp4 = []
tp4.append(tp[4])
tp5 = []
tp5.append(tp[5])
tp6 = []
tp6.append(tp[6])
tp7 = []
tp7.append(tp[7])
for i in range(len(itog_e)):
    if i < number_of_products:
        tp4.append(df_3.loc[itog_e[i], cL_itogo])
    elif i < number_of_products * 2:
        tp5.append(df_3.loc[itog_e[i], cL_itogo])
    elif i < number_of_products * 3:
        tp6.append(df_3.loc[itog_e[i], cL_itogo])
    elif i < number_of_products * 4:                #Это если 4 ТП иначе уменьшать или увеличивать
        tp7.append(df_3.loc[itog_e[i], cL_itogo])


del tp0[0] #удаляем Имя ТП из списка
del tp1[0]
del tp2[0]
del tp3[0]
del tp4[0]
del tp5[0]
del tp6[0]
del tp7[0]

df_4 = pd.DataFrame([tp0, tp1, tp2, tp3, tp4, tp5, tp6, tp7],
                    columns=columns,
                    index=tp)
#print(df_4)

group_koef = []
group_koef.extend(d)
group_koef.extend(d_s)
group_koef.extend(v)
group_koef.extend(v_s)
group_koef.extend(v_o_s)
group_koef.extend(m)
group_koef.extend(k)
group_koef.extend(k_s)
group_koef.extend(t)

# group_pvz = [] не нужно раскоменчивать
# group_pvz.extend(Dolgacheva)
# group_pvz.extend(Drozdova)
# group_pvz.extend(Korsakova)
# group_pvz.extend(Krasnova)
# group_pvz.extend(Kudrova)
# group_pvz.extend(Migushova)
# group_pvz.extend(Fedotova)
# group_pvz.extend(Chepa)


for i, row in df.iterrows():
    #if df.loc[i, 3] in Dolgacheva: df.loc[i, 1] = tp[0] #убрал 1и2 If, сместил индексы в tp -1
    #if df.loc[i, 1] == tp_1c[0]: df.loc[i, 1] = tp[0]
    if df.loc[i, 3] in Drozdova: df.loc[i, 1] = tp[0]
    if df.loc[i, 1] == tp_1c[0]: df.loc[i, 1] = tp[0]
    if df.loc[i, 3] in Korsakova: df.loc[i, 1] = tp[1]
    if df.loc[i, 1] == tp_1c[1]: df.loc[i, 1] = tp[1]
    if df.loc[i, 3] in Krasnova: df.loc[i, 1] = tp[4]
    if df.loc[i, 1] == tp_1c[4]: df.loc[i, 1] = tp[4]
    if df.loc[i, 3] in Shutova: df.loc[i, 1] = tp[7]
    if df.loc[i, 1] == tp_1c[7]: df.loc[i, 1] = tp[7]
    if df.loc[i, 3] in Migushova: df.loc[i, 1] = tp[5]
    if df.loc[i, 1] == tp_1c[5]: df.loc[i, 1] = tp[5]
    if df.loc[i, 3] in Fedotova: df.loc[i, 1] = tp[3]
    if df.loc[i, 1] == tp_1c[3]: df.loc[i, 1] = tp[3]
    if df.loc[i, 3] in Chepa: df.loc[i, 1] = tp[6]
    if df.loc[i, 1] == tp_1c[6]: df.loc[i, 1] = tp[6]
    if df.loc[i, 3] in Ogurova: df.loc[i, 1] = tp[2]
    if df.loc[i, 1] == tp_1c[2]: df.loc[i, 1] = tp[2]
    for f in range(len(group_koef)):
        if df.loc[i, 4] == group_koef[f]: #Можно искать по названию в [i, 5], тогда f=1 в начале
            df.loc[i, 8] = group_koef[f+2] # Тогда тут f+1
            df.loc[i, 9] = df.loc[i, 6] / df.loc[i, 8]
            if df.loc[i, 4] in d: df.loc[i, 10] = '1_Премиум' #тут везде ставим поиск тогда в [i, 5]
            if df.loc[i, 4] in d_s: df.loc[i, 10] = '2_Премиум соль'
            if df.loc[i, 4] in v: df.loc[i, 10] = '3_Полосатая'
            if df.loc[i, 4] in v_s: df.loc[i, 10] = '4_Полосатая соль'
            if df.loc[i, 4] in v_o_s: df.loc[i, 10] = '5_Полосатая особо соль'
            if df.loc[i, 4] in m: df.loc[i, 10] = '6_Мастер'
            if df.loc[i, 4] in k: df.loc[i, 10] = '8_Караван'
            if df.loc[i, 4] in k_s: df.loc[i, 10] = '9_Караван стандарт'
            if df.loc[i, 4] in t: df.loc[i, 10] = '7_Тыква'
        else:
            f = f+3

products_in_svod = ['1_Премиум', '2_Премиум соль', '3_Полосатая', '4_Полосатая соль', '5_Полосатая особо соль',
                    '7_Тыква', '8_Караван', '9_Караван стандарт']

#    for pvz in range(len(group_pvz)):  не раскомменчивать
#        if df.loc[i, 3] in Dolgacheva: df.loc[i, 1] = 'Долгачева Ирина (ЭК СМАРТ'
#        if df.loc[i, 3] in Drozdova: df.loc[i, 1] = 'Дроздова Марина (ЭК СМАРТ'
#        if df.loc[i, 3] in Korsakova: df.loc[i, 1] = 'Корсакова Елена (ЭК СМАРТ'
#        if df.loc[i, 3] in Krasnova: df.loc[i, 1] = 'Краснова Наталья (ЭК СМАР'
#        if df.loc[i, 3] in Kudrova: df.loc[i, 1] = 'Кудрова Наталия (ЭК СМАРТ'
#        if df.loc[i, 3] in Migushova: df.loc[i, 1] = 'Мигушова Надежда (ЭК СМАР'
#        if df.loc[i, 3] in Fedotova: df.loc[i, 1] = 'Федотова Анна (ЭК СМАРТ)'
#        if df.loc[i, 3] in Chepa: df.loc[i, 1] = 'Чепа Елена (ЭК СМАРТ)'

agg_func_sum = {6:['sum']}
prom_df = df.groupby([1, 4, 5]).agg(agg_func_sum) #Тут тогда 4 надо исключать, будет только по 5

svod_df = pd.crosstab(df[1],
                      df[10],
                      values=df[9],
                      aggfunc='sum',
                      normalize=False)
svod_df = svod_df.round(2)
svod_df.index.name = None
svod_df.rename(columns={'1_Премиум': 'Премиум', '2_Премиум соль': 'Премиум соль', '3_Полосатая': 'Полосатая',
                        '4_Полосатая соль': 'Полосатая соль', '5_Полосатая особо соль': 'Полосатая особо соль',
                        '7_Тыква': 'Тыква', '8_Караван': 'Караван', '9_Караван стандарт': 'Караван стандарт'},
               inplace=True)

# print(svod_df.index)
# print(df_4.index)


del tp0[5], tp0[-1] #удаляем Мастер Жарки и Арахис Джинн из списка
del tp1[5], tp1[-1]
del tp2[5], tp2[-1]
del tp3[5], tp3[-1]
del tp4[5], tp4[-1]
del tp5[5], tp5[-1]
del tp6[5], tp6[-1]
del tp7[5], tp7[-1]
#test = tp0
del columns[5], columns[-1]
df_5 = pd.DataFrame([tp0, tp1, tp2, tp3, tp4, tp5, tp6, tp7],
                    columns=columns,
                    index=tp)

df_6 = svod_df.loc[tp]
#print(df_6)
df_7 = pd.DataFrame(df_6.values - df_5.values, df_5.index, columns)
# print(df_7)

# for i, row in svod_df.iterrows(): #Проверить т.к. ТП в своде вразнобой
#     if svod_df.loc[i, 1] == df_4.loc[0, 'ТП']:
#         b = 1
#         for b in range(len(tp0)):
#             tp0[b] = svod_df.loc[svod_df.index[i], products_in_svod[b-1]] - tp0[b]
#     if svod_df.index[i] == df_4.loc[1, 'ТП']:
#         b = 1
#         for b in range(len(tp1)):
#             tp1[b] = svod_df.loc[svod_df.index[i], products_in_svod[b-1]] - tp1[b]
#     if svod_df.index[i] == df_4.loc[2, 'ТП']:
#         b = 1
#         for b in range(len(tp2)):
#             tp2[b] = svod_df.loc[svod_df.index[i], products_in_svod[b-1]] - tp2[b]
#     if svod_df.index[i] == df_4.loc[3, 'ТП']:
#         b = 1
#         for b in range(len(tp3)):
#             tp3[b] = svod_df.loc[svod_df.index[i], products_in_svod[b-1]] - tp3[b]
#     if svod_df.index[i] == df_4.loc[4, 'ТП']:
#         b = 1
#         for b in range(len(tp4)):
#             tp4[b] = svod_df.loc[svod_df.index[i], products_in_svod[b - 1]] - tp4[b]
#     if svod_df.index[i] == df_4.loc[5, 'ТП']:
#         b = 1
#         for b in range(len(tp5)):
#             tp5[b] = svod_df.loc[svod_df.index[i], products_in_svod[b - 1]] - tp5[b]
#     if svod_df.index[i] == df_4.loc[6, 'ТП']:
#         b = 1
#         for b in range(len(tp6)):
#             tp6[b] = svod_df.loc[svod_df.index[i], products_in_svod[b - 1]] - tp6[b]
#     if svod_df.index[i] == df_4.loc[7, 'ТП']:
#         b = 1
#         for b in range(len(tp7)):
#             tp7[b] = svod_df.loc[svod_df.index[i], products_in_svod[b - 1]] - tp7[b]
#
# df_5 = pd.DataFrame([tp0, tp1, tp2, tp3, tp4, tp5, tp6, tp7],
#                     columns=columns)
# print(df_5)

#test = svod_df.loc[svod_df.index[0], products_in_svod[0]]
# test = df_4.loc[0, 'ТП']
# test = svod_df.index[0]


# new_index = ['Дроздова Марина', 'Корсакова Елена', 'Огурова Ольга', 'Федотова Анна',
#              'Краснова Наталья', 'Мигушова Надежда', 'Чепа Елена', 'Шутова Ольга', 'Сотрудники']
# svod_df.reindex(new_index)
# print(svod_df.loc['Дроздова Марина', '2_Премиум соль'])
# print(svod_df.index)
# print(svod_df)



# itog = []
# for i in range(len(itog_k)):
#     if itog.append(df_2.loc[i, cL_itogo])

# columns = ['ТП']
# columns.extend(products_groups_sv_k[:number_of_products])


# for i, row in svod_df.iterrows():
#     if svod_df.index[i] == 'Долгачева Ирина (ЭК СМАРТ': svod_df.set_index[i] = 'Долгачева Ирина'
#     if svod_df.index[i] == 'Дроздова Марина (ЭК СМАРТ': svod_df.set_index[i] = 'Дроздова Марина'
#     if svod_df.index[i] == 'Корсакова Елена (ЭК СМАРТ': svod_df.set_index[i] = 'Корсакова Елена'
#     if svod_df.index[i] == 'Огурова Ольга (ЭК СМАРТ)': svod_df.set_index[i] = 'Огурова Ольгаа'
#     if svod_df.index[i] == 'Федотова Анна (ЭК СМАРТ)': svod_df.set_index[i] = 'Федотова Анна'
#     if svod_df.index[i] == 'Краснова Наталья (ЭК СМАР': svod_df.set_index[i] = 'Краснова Наталья'
#     if svod_df.index[i] == 'Мигушова Надежда (ЭК СМАР': svod_df.set_index[i] = 'Мигушова Надежда'
#     if svod_df.index[i] == 'Чепа Елена (ЭК СМАРТ)': svod_df.set_index[i] = 'Чепа Елена'
#     if svod_df.index[i] == 'Шутова Ольга (ЭК СМАРТ)': svod_df.set_index[i] = 'Шутова Ольга'
# print(svod_df)
# clmns_svod = list(svod_df)
# for c in range(len(clmns_svod)):
#     if svod_df.loc[1, c] == '1_Премиум': svod_df.loc[1, c] = 'Премиум'
#     if svod_df.loc[1, c] == '2_Премиум соль': svod_df.loc[1, c] = 'Премиум соль'
#     if svod_df.loc[1, c] == '3_Полосатая': svod_df.loc[1, c] = 'Полосатая'
#     if svod_df.loc[1, c] == '4_Полосатая соль': svod_df.loc[1, c] = 'Полосатая соль'
#     if svod_df.loc[1, c] == '5_Полосатая особо соль': svod_df.loc[1, c] = 'Полосатая особо соль'
#     if svod_df.loc[1, c] == '6_Мастер': svod_df.loc[1, c] = 'Мастер Жарки'
#     if svod_df.loc[1, c] == '7_Тыква': svod_df.loc[1, c] = 'Тыква'
#     if svod_df.loc[1, c] == '8_Караван': svod_df.loc[1, c] = 'Караван'
#     if svod_df.loc[1, c] == '9_Караван стандарт': svod_df.loc[1, c] = 'Караван СТАНДАРТ'


# for p in range(16):
#     if df_2.loc[cL_Dolgacheva+p, 0] == 'Премиум': raznica = svod_df[- df_2.loc[cL_Dolgacheva+p, cL_itogo]


with pd.ExcelWriter(file_name_result, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
    df_6.to_excel(writer, sheet_name='Сводка 2.0', startrow=0)
    df_7.to_excel(writer, sheet_name='Корректировка')
    svod_df.to_excel(writer, sheet_name='Сводка', startrow=0)
    prom_df.to_excel(writer, sheet_name='Пачки')
    df.to_excel(writer, sheet_name='Полная')
    df_4.to_excel(writer, sheet_name='СВ')



